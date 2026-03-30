from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Optional

import httpx
from openpyxl import load_workbook
from sqlmodel import Session, select

from .models import (
    BookkeepingEntry,
    BookkeepingImport,
    DiscordMessage,
    Transaction,
    normalize_money_value,
    signed_money_delta,
    utcnow,
)
from .transactions import transaction_base_query
from .db import engine, managed_session


KIND_ALIASES = {
    "sale": "sale",
    "sell": "sale",
    "sold": "sale",
    "income": "sale",
    "revenue": "sale",
    "buy": "buy",
    "bought": "buy",
    "purchase": "buy",
    "inventory": "buy",
    "trade": "trade",
    "expense": "expense",
    "expenses": "expense",
}

PAYMENT_ALIASES = {
    "cash": "cash",
    "zelle": "zelle",
    "venmo": "venmo",
    "paypal": "paypal",
    "card": "card",
    "tap": "card",
}

CATEGORY_TEXT_HINTS = {
    "inventory": ("inventory", "slab", "slabs", "single", "singles", "sealed", "binder", "collection"),
    "fees": ("fee", "fees", "vendor", "booth"),
    "travel": ("travel", "gas", "parking", "hotel", "uber", "lyft"),
    "food": ("food", "meal", "lunch", "dinner", "snack", "coffee"),
    "supplies": ("supply", "supplies", "toploader", "sleeve", "tape", "mailer"),
}

ENTRY_KIND_HINTS = {
    "sale": ("sale", "sell", "sold", "income", "revenue"),
    "buy": ("buy", "bought", "purchase", "inventory"),
    "trade": ("trade", "swap"),
    "expense": ("expense", "expenses", "fee", "supplies", "rent", "gas", "meal"),
}

DATE_HEADER_HINTS = ("date", "sold_at", "purchased_at", "occurred_at", "transaction_date")
ENTRY_KIND_HEADER_HINTS = ("entry_kind", "kind", "type", "transaction_type")
PAYMENT_HEADER_HINTS = ("payment", "payment_method", "method", "tender")
CATEGORY_HEADER_HINTS = ("category", "expense_category", "department")
NOTES_HEADER_HINTS = ("notes", "description", "memo", "details", "item", "items", "product")
MONEY_IN_HEADER_HINTS = ("money_in", "cash_in", "credit", "deposit", "income", "sale", "sales")
MONEY_OUT_HEADER_HINTS = ("money_out", "cash_out", "debit", "expense", "expenses", "buy", "buys", "purchase", "cost")
AMOUNT_HEADER_HINTS = ("amount", "total", "price", "cash", "net", "value")
SHEET_ENTRY_KIND_HINTS = {
    "income": "sale",
    "spending": "expense",
    "trade": "trade",
    "overview": None,
}


def extract_google_sheet_url(text: str) -> Optional[str]:
    match = re.search(r"https://docs\.google\.com/spreadsheets/[^\s>]+", text or "", re.I)
    return match.group(0) if match else None


def normalize_google_sheet_url(sheet_url: str) -> str:
    return (sheet_url or "").split("#", 1)[0].strip()


def extract_google_sheet_gid(sheet_url: str) -> Optional[str]:
    match = re.search(r"gid=(\d+)", sheet_url or "", re.I)
    return match.group(1) if match else None


def build_google_sheet_export_url(sheet_url: str) -> tuple[str, str]:
    normalized = normalize_google_sheet_url(sheet_url)
    normalized = re.sub(r"/edit(?:\?.*)?$", "", normalized)
    return f"{normalized}/export?format=xlsx", ".xlsx"


def infer_show_label_from_message(message_text: str, fallback_name: str) -> str:
    text = (message_text or "").strip()
    if ":" in text:
        prefix = text.split(":", 1)[0].strip()
        if prefix and len(prefix) <= 80:
            return prefix
    return fallback_name


def infer_show_date_from_text(message_text: str, fallback_year: int) -> Optional[datetime]:
    text = (message_text or "").strip()
    month_day_match = re.search(
        r"\b(jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|aug|august|sep|sept|september|oct|october|nov|november|dec|december)\s+(\d{1,2})\b",
        text,
        re.I,
    )
    if month_day_match:
        month = month_day_match.group(1)
        day = month_day_match.group(2)
        for fmt in ("%B %d %Y", "%b %d %Y"):
            try:
                return datetime.strptime(f"{month} {day} {fallback_year}", fmt)
            except ValueError:
                continue
    return None


def get_existing_import_by_source_url(session: Session, source_url: str) -> Optional[BookkeepingImport]:
    normalized = normalize_google_sheet_url(source_url)
    return session.exec(
        select(BookkeepingImport).where(BookkeepingImport.source_url == normalized)
    ).first()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def looks_like_header(header: str, hints: tuple[str, ...]) -> bool:
    return any(hint == header or hint in header for hint in hints)


def clean_text_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    return text or None


def infer_payment_method_from_text(text: Optional[str]) -> Optional[str]:
    normalized = clean_text_value(text)
    if not normalized:
        return None
    lower = normalized.lower()
    for token, payment_method in PAYMENT_ALIASES.items():
        if token in lower:
            return payment_method
    return None


def infer_category_from_text(text: Optional[str]) -> Optional[str]:
    normalized = clean_text_value(text)
    if not normalized:
        return None
    lower = normalized.lower()
    for category, hints in CATEGORY_TEXT_HINTS.items():
        if any(hint in lower for hint in hints):
            return category
    return None


def parse_optional_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    is_negative = "(" in text and ")" in text
    text = text.replace(",", "").replace("$", "").replace("(", "").replace(")", "")
    if text in {"-", "--"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    value = float(match.group(0))
    if is_negative and value > 0:
        value *= -1
    return value


def parse_strict_numeric_cell(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    is_negative = text.startswith("(") and text.endswith(")")
    normalized = text.replace(",", "").replace("$", "").replace("(", "").replace(")", "").strip()
    if normalized in {"", "-", "--"}:
        return None
    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return None

    parsed = float(normalized)
    if is_negative and parsed > 0:
        parsed *= -1
    return parsed


def parse_optional_datetime(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 60000:
        excel_epoch = datetime(1899, 12, 30)
        return excel_epoch + timedelta(days=float(value))
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    if not text:
        return None

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%y",
        "%Y/%m/%d",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%B %d %Y",
        "%b %d %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        if parsed.tzinfo is not None:
            return parsed.astimezone(timezone.utc).replace(tzinfo=None)
        return parsed
    except ValueError:
        return None


def read_tabular_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        return [
            {
                "__sheet_name": "import",
                **dict(row),
            }
            for row in reader
        ]

    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), data_only=True)
        parsed_rows: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header_row = next((row for row in rows if any(cell not in (None, "") for cell in row)), None)
            if not header_row:
                continue
            header_index = rows.index(header_row)
            headers = [normalize_header(cell) or f"column_{index+1}" for index, cell in enumerate(header_row)]
            for row in rows[header_index + 1:]:
                if not any(cell not in (None, "") for cell in row):
                    continue
                parsed_rows.append(
                    {
                        "__sheet_name": sheet.title,
                        **{
                            headers[index]: row[index] if index < len(row) else None
                            for index in range(len(headers))
                        },
                    }
                )
        return parsed_rows

    raise ValueError("Only .csv and .xlsx files are supported")


def infer_entry_kind(row: dict[str, Any]) -> Optional[str]:
    return infer_entry_kind_from_headers(row) or infer_entry_kind_from_values(row)


def infer_entry_kind_from_headers(row: dict[str, Any]) -> Optional[str]:
    for key, value in row.items():
        key_lower = key.lower()
        value_lower = str(value or "").strip().lower()
        if looks_like_header(key_lower, ENTRY_KIND_HEADER_HINTS):
            for token, normalized in KIND_ALIASES.items():
                if token in value_lower:
                    return normalized
    return None


def infer_entry_kind_from_values(row: dict[str, Any]) -> Optional[str]:
    for value in row.values():
        value_lower = str(value or "").strip().lower()
        for normalized, tokens in ENTRY_KIND_HINTS.items():
            if any(token in value_lower for token in tokens):
                return normalized
    return None


def infer_entry_kind_from_sheet_name(sheet_name: Optional[str]) -> Optional[str]:
    normalized = normalize_header(sheet_name)
    if not normalized:
        return None
    for token, entry_kind in SHEET_ENTRY_KIND_HINTS.items():
        if token in normalized:
            return entry_kind
    return None


def infer_payment_method(row: dict[str, Any]) -> Optional[str]:
    for key, value in row.items():
        key_lower = key.lower()
        value_lower = str(value or "").strip().lower()
        if looks_like_header(key_lower, PAYMENT_HEADER_HINTS):
            for token, normalized in PAYMENT_ALIASES.items():
                if token in value_lower:
                    return normalized
    for value in row.values():
        value_lower = str(value or "").strip().lower()
        for token, normalized in PAYMENT_ALIASES.items():
            if token in value_lower:
                return normalized
    return None


def infer_category(row: dict[str, Any]) -> Optional[str]:
    for key, value in row.items():
        key_lower = key.lower()
        if looks_like_header(key_lower, CATEGORY_HEADER_HINTS):
            text = clean_text_value(value)
            if text:
                return text.lower()
    return None


def infer_occurred_at(row: dict[str, Any]) -> Optional[datetime]:
    for key, value in row.items():
        if looks_like_header(key.lower(), DATE_HEADER_HINTS):
            parsed = parse_optional_datetime(value)
            if parsed:
                return parsed
    for value in row.values():
        parsed = parse_optional_datetime(value)
        if parsed:
            return parsed
    return None


def collect_amount_candidates(row: dict[str, Any], hints: tuple[str, ...]) -> list[float]:
    values: list[float] = []
    for key, value in row.items():
        if looks_like_header(key.lower(), hints):
            parsed = parse_optional_float(value)
            if parsed is not None:
                values.append(parsed)
    return values


def collect_generic_numeric_candidates(row: dict[str, Any]) -> list[float]:
    values: list[float] = []
    for key, value in row.items():
        if not re.fullmatch(r"column_\d+", key.lower()):
            continue
        parsed = parse_strict_numeric_cell(value)
        if parsed is not None:
            values.append(parsed)
    return values


def is_header_like_row(row: dict[str, Any]) -> bool:
    if not row:
        return False

    header_like_values = 0
    populated_values = 0
    known_header_tokens = {
        "category",
        "transaction name",
        "amount",
        "trade",
        "details",
        "notes",
        "payment",
        "payment method",
        "date",
        "type",
        "kind",
    }

    for value in row.values():
        text = clean_text_value(value)
        if not text:
            continue
        populated_values += 1
        if normalize_header(text).replace("_", " ") in known_header_tokens:
            header_like_values += 1

    return populated_values > 0 and header_like_values >= max(2, populated_values // 2)


def infer_amount_fields(row: dict[str, Any]) -> tuple[Optional[float], Optional[float], Optional[float]]:
    money_in_candidates = collect_amount_candidates(row, MONEY_IN_HEADER_HINTS)
    money_out_candidates = collect_amount_candidates(row, MONEY_OUT_HEADER_HINTS)
    amount_candidates = collect_amount_candidates(row, AMOUNT_HEADER_HINTS)
    generic_numeric_candidates = collect_generic_numeric_candidates(row)

    money_in = next((abs(value) for value in money_in_candidates if value != 0), None)
    money_out = next((abs(value) for value in money_out_candidates if value != 0), None)

    amount = None
    for candidate in amount_candidates:
        if candidate == 0:
            continue
        amount = abs(candidate)
        if candidate < 0 and money_out is None:
            money_out = abs(candidate)
        elif candidate > 0 and money_in is None:
            money_in = abs(candidate)
        break

    if amount is None:
        fallback_generic = next((value for value in generic_numeric_candidates if value not in (None, 0)), None)
        if fallback_generic is not None:
            amount = abs(fallback_generic)
            if fallback_generic < 0 and money_out is None:
                money_out = abs(fallback_generic)
            elif fallback_generic > 0 and money_in is None:
                money_in = abs(fallback_generic)

    if amount is None:
        fallback_candidates = []
        for key, value in row.items():
            key_lower = key.lower()
            if looks_like_header(
                key_lower,
                DATE_HEADER_HINTS + ENTRY_KIND_HEADER_HINTS + PAYMENT_HEADER_HINTS + CATEGORY_HEADER_HINTS + NOTES_HEADER_HINTS,
            ):
                continue
            fallback_candidates.append(parse_optional_float(value))
        fallback = next((value for value in fallback_candidates if value not in (None, 0)), None)
        if fallback is not None:
            amount = abs(fallback)
            if fallback < 0 and money_out is None:
                money_out = abs(fallback)
            elif fallback > 0 and money_in is None:
                money_in = abs(fallback)

    if amount is None:
        if money_in is not None and money_out is None:
            amount = money_in
        elif money_out is not None and money_in is None:
            amount = money_out
        elif money_in is not None and money_out is not None:
            amount = max(money_in, money_out)

    return amount, money_in, money_out


def infer_entry_kind_from_amounts(
    row: dict[str, Any],
    amount: Optional[float],
    money_in: Optional[float],
    money_out: Optional[float],
    sheet_name: Optional[str] = None,
) -> Optional[str]:
    inferred = infer_entry_kind_from_headers(row)
    if inferred:
        return inferred

    sheet_inferred = infer_entry_kind_from_sheet_name(sheet_name)
    if sheet_inferred:
        return sheet_inferred

    inferred = infer_entry_kind_from_values(row)
    if inferred:
        return inferred

    category = infer_category(row) or ""
    joined = " ".join(str(value or "").lower() for value in row.values())
    if money_in and not money_out:
        return "sale"
    if money_out and not money_in:
        if "inventory" in category or "inventory" in joined:
            return "buy"
        if any(token in joined for token in ("buy", "bought", "purchase")):
            return "buy"
        if any(token in joined for token in ("expense", "fee", "gas", "rent", "meal", "food", "supply")):
            return "expense"
        return "expense"
    if amount and "trade" in joined:
        return "trade"
    return None


def infer_notes(row: dict[str, Any]) -> Optional[str]:
    preferred_parts = []
    extra_parts = []
    for key, value in row.items():
        cleaned = clean_text_value(value)
        if cleaned is None:
            continue
        key_lower = key.lower()
        if looks_like_header(key_lower, DATE_HEADER_HINTS + ENTRY_KIND_HEADER_HINTS + PAYMENT_HEADER_HINTS + CATEGORY_HEADER_HINTS):
            continue
        if looks_like_header(key_lower, MONEY_IN_HEADER_HINTS + MONEY_OUT_HEADER_HINTS + AMOUNT_HEADER_HINTS):
            continue
        part = f"{key}: {cleaned}"
        if looks_like_header(key_lower, NOTES_HEADER_HINTS):
            preferred_parts.append(part)
        else:
            extra_parts.append(part)
    parts = preferred_parts + extra_parts
    return " | ".join(parts) or None


def normalize_bookkeeping_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    for index, row in enumerate(rows, start=1):
        sheet_name = clean_text_value(row.get("__sheet_name"))
        sheet_default_entry_kind = infer_entry_kind_from_sheet_name(sheet_name)
        if sheet_default_entry_kind is None and normalize_header(sheet_name) == "overview":
            continue

        normalized_row = {
            normalize_header(key): value
            for key, value in row.items()
            if key != "__sheet_name" and normalize_header(key)
        }
        if not normalized_row:
            continue
        if is_header_like_row(normalized_row):
            continue

        date_value = infer_occurred_at(normalized_row)
        amount_value, money_in, money_out = infer_amount_fields(normalized_row)
        entry_kind = infer_entry_kind_from_amounts(
            normalized_row,
            amount_value,
            money_in,
            money_out,
            sheet_name=sheet_name,
        )
        notes = infer_notes(normalized_row)
        payment_method = (
            infer_payment_method(normalized_row)
            or infer_payment_method_from_text(notes)
            or infer_payment_method_from_text(sheet_name)
        )
        category = (
            infer_category(normalized_row)
            or infer_category_from_text(notes)
            or infer_category_from_text(sheet_name)
        )

        if amount_value is None and not notes:
            continue

        normalized.append(
            {
                "row_index": index,
                "sheet_name": sheet_name,
                "occurred_at": date_value,
                "entry_kind": entry_kind,
                "amount": normalize_money_value(amount_value) if amount_value is not None else None,
                "payment_method": payment_method,
                "category": category,
                "notes": notes,
                "raw_row_json": json.dumps(normalized_row, default=str),
            }
        )
    return normalized


def import_bookkeeping_file(
    session: Session,
    *,
    filename: str,
    content: bytes,
    show_label: str,
    show_date: Optional[datetime],
    range_start: Optional[datetime],
    range_end: Optional[datetime],
    source_url: Optional[str] = None,
    source_kind: str = "upload",
) -> BookkeepingImport:
    normalized_source_url = normalize_google_sheet_url(source_url) if source_url else None
    if normalized_source_url:
        existing = get_existing_import_by_source_url(session, normalized_source_url)
        if existing:
            return existing

    raw_rows = read_tabular_rows(filename, content)
    normalized_rows = normalize_bookkeeping_rows(raw_rows)

    bookkeeping_import = BookkeepingImport(
        show_label=show_label,
        show_date=show_date,
        range_start=range_start,
        range_end=range_end,
        source_kind=source_kind,
        source_name=filename,
        source_url=normalized_source_url,
        row_count=len(normalized_rows),
    )
    session.add(bookkeeping_import)
    session.commit()
    session.refresh(bookkeeping_import)

    for row in normalized_rows:
        session.add(
            BookkeepingEntry(
                import_id=bookkeeping_import.id,
                row_index=row["row_index"],
                sheet_name=row["sheet_name"],
                occurred_at=row["occurred_at"],
                entry_kind=row["entry_kind"],
                amount=row["amount"],
                payment_method=row["payment_method"],
                category=row["category"],
                notes=row["notes"],
                raw_row_json=row["raw_row_json"],
            )
        )
    session.commit()
    return bookkeeping_import


async def auto_import_public_google_sheet(
    *,
    message_text: str,
    created_at: Optional[datetime],
    sheet_url: str,
) -> Optional[int]:
    export_url, suffix = build_google_sheet_export_url(sheet_url)
    show_label = infer_show_label_from_message(message_text, "Discord bookkeeping import")
    show_date = infer_show_date_from_text(message_text, (created_at or utcnow()).year)
    range_start = show_date.replace(hour=0, minute=0, second=0, microsecond=0) if show_date else None
    range_end = show_date.replace(hour=23, minute=59, second=59, microsecond=999999) if show_date else None

    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        response = await client.get(export_url)
        response.raise_for_status()
        content = response.content

    filename = f"{re.sub(r'[^a-zA-Z0-9._-]+', '-', show_label).strip('-') or 'bookkeeping-import'}{suffix}"

    with managed_session() as session:
        imported = import_bookkeeping_file(
            session,
            filename=filename,
            content=content,
            show_label=show_label,
            show_date=show_date,
            range_start=range_start,
            range_end=range_end,
            source_url=sheet_url,
            source_kind="google_sheet_auto",
        )
        reconcile_bookkeeping_import(session, imported.id)
        return imported.id


async def refresh_bookkeeping_import_from_source(bookkeeping_import_id: int) -> Optional[int]:
    with managed_session() as session:
        existing = session.get(BookkeepingImport, bookkeeping_import_id)
        if not existing:
            raise ValueError("Bookkeeping import not found")
        if not existing.source_url:
            raise ValueError("This import has no source URL to refresh from")

        source_url = existing.source_url
        show_label = existing.show_label
        show_date = existing.show_date
        range_start = existing.range_start
        range_end = existing.range_end
        source_kind = existing.source_kind or "upload"
        source_name = existing.source_name or "bookkeeping-import.xlsx"

        rows = session.exec(
            select(BookkeepingEntry).where(BookkeepingEntry.import_id == bookkeeping_import_id)
        ).all()
        for row in rows:
            session.delete(row)
        session.delete(existing)
        session.commit()

    export_url, suffix = build_google_sheet_export_url(source_url)
    async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
        response = await client.get(export_url)
        response.raise_for_status()
        content = response.content

    filename = source_name
    if not filename.lower().endswith((".xlsx", ".csv")):
        filename = f"{filename}{suffix}"

    with managed_session() as session:
        imported = import_bookkeeping_file(
            session,
            filename=filename,
            content=content,
            show_label=show_label,
            show_date=show_date,
            range_start=range_start,
            range_end=range_end,
            source_url=source_url,
            source_kind=source_kind,
        )
        reconcile_bookkeeping_import(session, imported.id)
        return imported.id


def bookkeeping_entry_amount(entry: BookkeepingEntry) -> float:
    return normalize_money_value(entry.amount)


def transaction_match_amount(row: Transaction) -> float:
    if row.amount is not None:
        return normalize_money_value(row.amount)
    return max(normalize_money_value(row.money_in), normalize_money_value(row.money_out))


def normalize_comparison_text(value: Optional[str]) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def comparison_tokens(value: Optional[str]) -> set[str]:
    return {
        token
        for token in normalize_comparison_text(value).split()
        if len(token) >= 4 and token not in {"with", "from", "this", "that"}
    }


def date_distance_days(left: Optional[datetime], right: Optional[datetime]) -> Optional[float]:
    if not left or not right:
        return None
    return abs((left - right).total_seconds()) / 86400.0


def transaction_match_score(entry: BookkeepingEntry, row: Transaction) -> tuple[float, str]:
    score = 0.0
    reasons: list[str] = []
    sheet_name = entry.sheet_name or ""
    sheet_entry_kind = infer_entry_kind_from_sheet_name(sheet_name)
    sheet_payment_method = infer_payment_method_from_text(sheet_name)
    sheet_category = infer_category_from_text(sheet_name)

    entry_amount = bookkeeping_entry_amount(entry)
    tx_amount = transaction_match_amount(row)
    if entry.amount is not None and tx_amount:
        amount_delta = abs(entry_amount - tx_amount)
        if amount_delta < 0.001:
            score += 10.0
            reasons.append("exact_amount")
        elif amount_delta <= 1.0:
            score += 7.0
            reasons.append("close_amount")
        elif amount_delta <= 5.0:
            score += 4.0
            reasons.append("near_amount")
        else:
            score -= min(amount_delta, 25.0)
    elif entry.amount is not None:
        score -= 5.0

    if entry.entry_kind and row.entry_kind == entry.entry_kind:
        score += 5.0
        reasons.append("entry_kind")
    elif entry.entry_kind and row.entry_kind:
        score -= 2.5
    elif sheet_entry_kind and row.entry_kind == sheet_entry_kind:
        score += 2.0
        reasons.append("sheet_entry_kind")

    if entry.payment_method and row.payment_method == entry.payment_method:
        score += 2.0
        reasons.append("payment")
    elif entry.payment_method and row.payment_method:
        score -= 0.5
    elif sheet_payment_method and row.payment_method == sheet_payment_method:
        score += 1.0
        reasons.append("sheet_payment")

    entry_category = normalize_comparison_text(entry.category or sheet_category)
    tx_category = normalize_comparison_text(row.category or row.expense_category)
    if entry_category and tx_category:
        if entry_category == tx_category:
            score += 2.0
            reasons.append("category")
        elif entry_category in tx_category or tx_category in entry_category:
            score += 1.0
            reasons.append("category_partial")

    distance_days = date_distance_days(entry.occurred_at, row.occurred_at)
    if distance_days is not None:
        if distance_days <= 0.1:
            score += 3.0
            reasons.append("same_day")
        elif distance_days <= 1.0:
            score += 1.5
            reasons.append("near_day")
        elif distance_days <= 3.0:
            score += 0.5
        else:
            score -= min(distance_days, 7.0)

    entry_match_text = " ".join(
        part for part in [entry.sheet_name, entry.payment_method, entry.category, entry.notes] if part
    )
    tx_match_text = " ".join(
        part
        for part in [
            row.channel_name,
            row.payment_method,
            row.category,
            row.expense_category,
            row.notes,
            row.trade_summary,
            row.source_content,
        ]
        if part
    )
    overlapping_tokens = comparison_tokens(entry_match_text) & comparison_tokens(tx_match_text)
    if overlapping_tokens:
        score += min(len(overlapping_tokens), 4)
        reasons.append("notes")

    sheet_tokens = comparison_tokens(entry.sheet_name)
    channel_tokens = comparison_tokens(row.channel_name)
    if sheet_tokens and channel_tokens:
        overlap = sheet_tokens & channel_tokens
        if overlap:
            score += min(len(overlap), 2)
            reasons.append("sheet_channel")

    return score, ",".join(reasons)


def reconcile_match_status(entry: BookkeepingEntry, row: Transaction, score: float) -> str:
    entry_amount = bookkeeping_entry_amount(entry)
    tx_amount = transaction_match_amount(row)
    amount_exact = entry.amount is not None and abs(entry_amount - tx_amount) < 0.001
    kind_exact = bool(entry.entry_kind and row.entry_kind == entry.entry_kind)
    payment_exact = bool(entry.payment_method and row.payment_method == entry.payment_method)
    category_exact = bool(
        entry.category
        and normalize_comparison_text(entry.category)
        and normalize_comparison_text(entry.category) == normalize_comparison_text(row.category or row.expense_category)
    )

    if amount_exact and (kind_exact or payment_exact or category_exact or score >= 14):
        return "matched_exact"
    if amount_exact or score >= 8:
        return "matched_amount_only"
    return "unmatched"


def reconcile_bookkeeping_import(session: Session, bookkeeping_import_id: int) -> dict[str, Any]:
    import_row = session.get(BookkeepingImport, bookkeeping_import_id)
    if not import_row:
        raise ValueError("Bookkeeping import not found")

    entries = session.exec(
        select(BookkeepingEntry)
        .where(BookkeepingEntry.import_id == bookkeeping_import_id)
        .order_by(BookkeepingEntry.row_index)
    ).all()

    inferred_start = import_row.range_start
    inferred_end = import_row.range_end
    entry_dates = [entry.occurred_at for entry in entries if entry.occurred_at]
    if not inferred_start and entry_dates:
        inferred_start = min(entry_dates) - timedelta(days=1)
    if not inferred_end and entry_dates:
        inferred_end = max(entry_dates) + timedelta(days=1)

    transactions = session.exec(
        transaction_base_query(start=inferred_start, end=inferred_end)
    ).all()

    unused_transaction_ids = {row.id for row in transactions if row.id is not None}
    transactions_by_id = {row.id: row for row in transactions if row.id is not None}

    matched = 0
    mismatched = 0
    unmatched = 0
    detailed_rows = []

    for entry in entries:
        best_match: Optional[Transaction] = None
        best_status = "unmatched"
        best_score = float("-inf")
        best_reason = ""

        candidates = [row for row in transactions if row.id in unused_transaction_ids]
        for candidate in candidates:
            score, reason = transaction_match_score(entry, candidate)
            if score > best_score:
                best_match = candidate
                best_score = score
                best_reason = reason

        if best_match is not None:
            best_status = reconcile_match_status(entry, best_match, best_score)
            if best_status == "unmatched":
                best_match = None

        entry.matched_transaction_id = best_match.id if best_match else None
        entry.match_status = best_status
        session.add(entry)

        if best_match and best_match.id in unused_transaction_ids:
            unused_transaction_ids.remove(best_match.id)

        if best_status == "matched_exact":
            matched += 1
        elif best_status == "matched_amount_only":
            mismatched += 1
        else:
            unmatched += 1

        detailed_rows.append(
            {
                "row_index": entry.row_index,
                "sheet_name": entry.sheet_name or "",
                "occurred_at": entry.occurred_at.isoformat(sep=" ", timespec="seconds") if entry.occurred_at else "",
                "entry_kind": entry.entry_kind or "",
                "amount": entry.amount,
                "payment_method": entry.payment_method or "",
                "category": entry.category or "",
                "notes": entry.notes or "",
                "match_status": best_status,
                "match_reason": best_reason,
                "match_score": round(best_score, 2) if best_score != float("-inf") else None,
                "matched_transaction": best_match,
            }
        )

    session.commit()

    unmatched_transactions = [transactions_by_id[tx_id] for tx_id in unused_transaction_ids]

    imported_totals = Counter()
    for entry in entries:
        if entry.entry_kind and entry.amount is not None:
            imported_totals[entry.entry_kind] += normalize_money_value(entry.amount)

    parsed_totals = Counter()
    for row in transactions:
        if row.entry_kind:
            parsed_totals[row.entry_kind] += transaction_match_amount(row)

    return {
        "import": import_row,
        "entries": detailed_rows,
        "summary": {
            "import_rows": len(entries),
            "matched_exact": matched,
            "matched_amount_only": mismatched,
            "unmatched_rows": unmatched,
            "unmatched_transactions": len(unmatched_transactions),
            "matched_rows": matched + mismatched,
            "match_rate": round(((matched + mismatched) / len(entries)) * 100, 1) if entries else 0.0,
            "imported_totals": dict(imported_totals),
            "parsed_totals": dict(parsed_totals),
            "import_net": round(sum(
                normalize_money_value(entry.amount or 0.0) * (
                    1 if entry.entry_kind in {"sale", "trade"} else -1 if entry.entry_kind in {"buy", "expense"} else 0
                )
                for entry in entries
            ), 2),
            "parsed_net": round(sum(
                signed_money_delta(row.money_in, row.money_out) for row in transactions
            ), 2),
        },
        "unmatched_transactions": unmatched_transactions[:100],
    }


def list_bookkeeping_imports(session: Session) -> list[BookkeepingImport]:
    return session.exec(
        select(BookkeepingImport).order_by(
            BookkeepingImport.show_date.desc(),
            BookkeepingImport.created_at.desc(),
        )
    ).all()


def list_detected_bookkeeping_posts(session: Session, limit: int = 20) -> list[dict[str, Any]]:
    rows = session.exec(
        select(DiscordMessage)
        .where(DiscordMessage.content.contains("docs.google.com/spreadsheets"))
        .order_by(DiscordMessage.created_at.desc())
        .limit(limit)
    ).all()
    detected = []
    for row in rows:
        sheet_url = extract_google_sheet_url(row.content or "") or ""
        existing_import = get_existing_import_by_source_url(session, sheet_url) if sheet_url else None
        detected.append(
            {
                "message_id": row.id,
                "created_at": row.created_at.isoformat(sep=" ", timespec="seconds") if row.created_at else "",
                "channel_name": row.channel_name or "",
                "author_name": row.author_name or "",
                "content": row.content or "",
                "sheet_url": sheet_url,
                "import_id": existing_import.id if existing_import else None,
            }
        )
    return detected


def get_bookkeeping_status_by_message_ids(session: Session, message_ids: list[int]) -> dict[int, dict[str, Any]]:
    if not message_ids:
        return {}

    transactions = session.exec(
        select(Transaction).where(Transaction.source_message_id.in_(message_ids))
    ).all()
    transaction_by_source_message_id = {
        transaction.source_message_id: transaction
        for transaction in transactions
        if transaction.id is not None
    }
    transaction_ids = [transaction.id for transaction in transactions if transaction.id is not None]
    if not transaction_ids:
        return {}

    entries = session.exec(
        select(BookkeepingEntry)
        .where(BookkeepingEntry.matched_transaction_id.in_(transaction_ids))
        .order_by(BookkeepingEntry.created_at.desc())
    ).all()

    by_transaction_id: dict[int, BookkeepingEntry] = {}
    status_rank = {"matched_exact": 3, "matched_amount_only": 2, "unmatched": 1, None: 0}
    for entry in entries:
        if entry.matched_transaction_id is None:
            continue
        existing = by_transaction_id.get(entry.matched_transaction_id)
        if existing is None or status_rank.get(entry.match_status, 0) > status_rank.get(existing.match_status, 0):
            by_transaction_id[entry.matched_transaction_id] = entry

    result: dict[int, dict[str, Any]] = {}
    for source_message_id, transaction in transaction_by_source_message_id.items():
        matched_entry = by_transaction_id.get(transaction.id)
        if not matched_entry:
            result[source_message_id] = {
                "status": "unmatched",
                "label": "Unmatched",
                "sheet_name": "",
            }
            continue
        result[source_message_id] = {
            "status": matched_entry.match_status or "unmatched",
            "label": (
                "Matched"
                if matched_entry.match_status == "matched_exact"
                else "Partial"
                if matched_entry.match_status == "matched_amount_only"
                else "Unmatched"
            ),
            "sheet_name": matched_entry.sheet_name or "",
        }
    return result
